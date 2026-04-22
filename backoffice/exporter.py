import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

import db

BASE = Path(__file__).parent
TEMPLATE_PATH = BASE / "data" / "template.xlsx"

SHEET_IG_POSTS = "인스타그램 게시물 데이터"
SHEET_TT_POSTS = "틱톡 게시물 데이터"
SHEET_IG_FOLLOWERS = "인스타그램 팔로워수"
SHEET_TT_FOLLOWERS = "틱톡 팔로워수"

POSTS_DATA_START_ROW = 5
FOLLOWERS_DATA_START_ROW = 8
FOLLOWERS_MARKER = "기간 총 증가"


def template_exists() -> bool:
    return TEMPLATE_PATH.exists()


def _parse_dt(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s)
    except (ValueError, TypeError):
        return s


def _short(text, n=50):
    if not text:
        return ""
    first_line = str(text).splitlines()[0] if text else ""
    return first_line[:n] + ("…" if len(first_line) > n else "")


def _clear_rows(ws, start_row, end_row, min_col, max_col):
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.hyperlink = None


def _write(ws, row, col, value):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return None
    cell.value = value
    return cell


def _set_link(ws, row, col, url):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = "바로가기"
    if url:
        cell.hyperlink = url
        cell.style = "Hyperlink"


def _fill_ig_posts(ws):
    end_row = max(ws.max_row, POSTS_DATA_START_ROW)
    _clear_rows(ws, POSTS_DATA_START_ROW, end_row, 2, 9)
    posts = sorted(db.fetch_ig_posts(), key=lambda r: r["posted_at"] or "")
    for i, p in enumerate(posts):
        r = POSTS_DATA_START_ROW + i
        _write(ws, r, 2, _parse_dt(p["posted_at"]))
        _write(ws, r, 3, _short(p["caption"]))
        _set_link(ws, r, 4, p["link"])
        _write(ws, r, 5, p["reach"] or 0)
        _write(ws, r, 6, p["likes"] or 0)
        _write(ws, r, 7, p["comments"] or 0)
        _write(ws, r, 8, p["saves"] or 0)
        _write(ws, r, 9, (p["likes"] or 0) + (p["comments"] or 0) + (p["saves"] or 0))


def _fill_tt_posts(ws):
    end_row = max(ws.max_row, POSTS_DATA_START_ROW)
    _clear_rows(ws, POSTS_DATA_START_ROW, end_row, 2, 10)
    posts = sorted(db.fetch_tt_posts(), key=lambda r: r["posted_at"] or "")
    for i, p in enumerate(posts):
        r = POSTS_DATA_START_ROW + i
        _write(ws, r, 2, _parse_dt(p["posted_at"]))
        _write(ws, r, 3, _short(p["title"]))
        _set_link(ws, r, 4, p["link"])
        _write(ws, r, 5, p["views"] or 0)
        _write(ws, r, 6, p["likes"] or 0)
        _write(ws, r, 7, p["comments"] or 0)
        _write(ws, r, 8, p["saves"] or 0)
        _write(ws, r, 9, p["shares"] or 0)
        _write(ws, r, 10, (p["likes"] or 0) + (p["comments"] or 0) + (p["saves"] or 0) + (p["shares"] or 0))


def _find_marker_row(ws, marker, col=2, search_from=1, search_to=200):
    for row_i in range(search_from, min(ws.max_row, search_to) + 1):
        if ws.cell(row_i, col).value == marker:
            return row_i
    return None


def _fill_daily_followers(ws, platform):
    records = [r for r in db.fetch_followers(platform) if r.get("new_followers") is not None or r.get("total_followers") is not None]
    records.sort(key=lambda r: r["date"])
    marker_row = _find_marker_row(ws, FOLLOWERS_MARKER)
    end_row = (marker_row - 1) if marker_row else max(ws.max_row, FOLLOWERS_DATA_START_ROW + 30)
    _clear_rows(ws, FOLLOWERS_DATA_START_ROW, end_row, 2, 3)
    for i, rec in enumerate(records):
        r = FOLLOWERS_DATA_START_ROW + i
        try:
            dt = datetime.fromisoformat(rec["date"])
        except Exception:
            dt = rec["date"]
        _write(ws, r, 2, dt)
        _write(ws, r, 3, rec.get("new_followers"))

    total = db.latest_total(platform)
    if total is not None:
        _write(ws, 4, 4, total)


def _fill_ig_demographics(ws):
    demo = db.latest_ig_demographics()
    if not demo:
        return

    ag_row = _find_marker_row(ws, "성별 및 연령별 팔로워")
    cities_row = _find_marker_row(ws, "상위 도시")
    countries_row = _find_marker_row(ws, "상위 국가", col=5)

    if demo.get("total_followers") is not None:
        _write(ws, 4, 4, demo["total_followers"])

    if ag_row:
        age_gender = json.loads(demo.get("age_gender_json") or "[]")
        header_row = ag_row + 1
        _clear_rows(ws, header_row + 1, header_row + 7, 2, 4)
        for i, a in enumerate(age_gender):
            r = header_row + 1 + i
            _write(ws, r, 2, a.get("age"))
            _write(ws, r, 3, a.get("female"))
            _write(ws, r, 4, a.get("male"))

    top_cities = json.loads(demo.get("top_cities_json") or "[]")
    top_countries = json.loads(demo.get("top_countries_json") or "[]")
    if cities_row:
        header = cities_row + 1
        _clear_rows(ws, header + 1, header + 12, 2, 3)
        for i, c in enumerate(top_cities[:10]):
            r = header + 1 + i
            _write(ws, r, 2, c.get("city"))
            _write(ws, r, 3, c.get("share"))
    if countries_row:
        header = countries_row + 1
        _clear_rows(ws, header + 1, header + 12, 5, 6)
        for i, c in enumerate(top_countries[:10]):
            r = header + 1 + i
            _write(ws, r, 5, c.get("country"))
            _write(ws, r, 6, c.get("share"))


def build_xlsx(output_path: Path) -> Path:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}. Upload it in the Settings tab.")
    wb = load_workbook(TEMPLATE_PATH)
    _fill_ig_posts(wb[SHEET_IG_POSTS])
    _fill_tt_posts(wb[SHEET_TT_POSTS])
    _fill_daily_followers(wb[SHEET_TT_FOLLOWERS], "tiktok")
    _fill_daily_followers(wb[SHEET_IG_FOLLOWERS], "instagram")
    _fill_ig_demographics(wb[SHEET_IG_FOLLOWERS])
    wb.save(output_path)
    return output_path
